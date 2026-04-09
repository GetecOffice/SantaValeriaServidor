from django.http import HttpResponse
from openpyxl import Workbook
from django.db import connection
from django.utils.timezone import localtime


def ExportarServidoCorralExcel(request):

    fechainicio = request.POST.get('fecha1')
    fechafinal = request.POST.get('fecha2')

    consulta_sql = """
            SELECT  r.Folio, r.Porcentaje, c.Descripcion AS Corral, p.Descripcion AS Producto, r.CantidadAnimales, 
            r.CantidadSolicitada, r.Cantidad1, r.Cantidad2, (r.Cantidad1 + r.Cantidad2) AS TotalCantidad,
            ROUND((r.Cantidad1 + r.Cantidad2) / NULLIF(r.CantidadAnimales, 0), 2) AS PromedioPorAnimal, r.FechaSol, r.FechaServida1, r.FechaServida2
            FROM Aplicacion_tblrepartidor r
            LEFT JOIN Aplicacion_tblcorrales c ON r.IDCorral_id = c.ID
            LEFT JOIN Aplicacion_tblproductos p ON r.IDProducto_id = p.ID
            LEFT JOIN Aplicacion_tblestatus e ON r.IDEstatus_id = e.ID
            WHERE r.IDEstatus_id IN (10,11)
        AND DATE(r.FechaServida2) BETWEEN %s AND %s
    """

    with connection.cursor() as cursor:
        cursor.execute(consulta_sql, [fechainicio, fechafinal])
        TServidos = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Servido Corral"

    headers = [
        "Folio",
        "Porcentaje",
        "Corral",
        "Producto",
        "Cabezas por corral",
        "Kilos solicitados",
        "Servido mañana",
        "Servido tarde",
        "Total servido",
        "Kilos por cabeza",
        "Fecha Solicitada",
        "Fecha Servida Desayuno",
        "Fecha Servida Cena"
    ]

    ws.append(headers)

    for row in TServidos:

        fecha = row[10]
        fecha_servida1 = row[11]
        fecha_servida2 = row[12]

        ws.append([
            row[0],
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
            row[6],
            row[7],
            row[8],
            row[9],
            fecha,
            fecha_servida1,
            fecha_servida2
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[10].number_format = 'DD/MM/YYYY HH:MM'
        row[11].number_format = 'DD/MM/YYYY HH:MM'
        row[12].number_format = 'DD/MM/YYYY HH:MM'

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ServidoCorral.xlsx'

    wb.save(response)

    return response