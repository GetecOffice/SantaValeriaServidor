from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models import Q
from django.utils.timezone import localtime
from Aplicacion.models import tblRepartidor

def ExportarServidoCorralExcel(request):

    TServidos = tblRepartidor.objects.filter(
        Q(IDEstatus_id=10) | Q(IDEstatus_id=11)
    ).values(
        'ID',
        'Folio',
        'IDCorral_id__Descripcion',
        'IDProducto_id__Descripcion',
        'IDEstatus_id__Descripcion',
        'CantidadSolicitada',
        'CantidadServida',
        'Fecha',
        'FechaServida'
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Servido Corral"

    headers = [
        "ID",
        "Folio",
        "Corral",
        "Producto",
        "Estatus",
        "Cantidad Solicitada",
        "Cantidad Servida",
        "Fecha",
        "Fecha Servida"
    ]

    ws.append(headers)

    # 🔹 Agregar datos
    for row in TServidos:

        fecha = row['Fecha']
        fecha_servida = row['FechaServida']

        if fecha:
            fecha = localtime(fecha).replace(tzinfo=None)

        if fecha_servida:
            fecha_servida = localtime(fecha_servida).replace(tzinfo=None)

        ws.append([
            row['ID'],
            row['Folio'],
            row['IDCorral_id__Descripcion'],
            row['IDProducto_id__Descripcion'],
            row['IDEstatus_id__Descripcion'],
            row['CantidadSolicitada'],
            row['CantidadServida'],
            fecha,
            fecha_servida,
        ])

    # 🔹 Aplicar formato de fecha (DESPUÉS de insertar datos)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[7].number_format = 'DD/MM/YYYY HH:MM'
        row[8].number_format = 'DD/MM/YYYY HH:MM'

    # 🔹 Ajustar ancho automático
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        ws.column_dimensions[column_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ServidoCorral.xlsx'

    wb.save(response)

    return response